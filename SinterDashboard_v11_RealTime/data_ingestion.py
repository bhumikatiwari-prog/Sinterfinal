"""
data_ingestion.py
─────────────────────────────────────────────────────────────
Automated Excel Data Aggregation & Processing Module
Handles upload → validation → daily aggregation → pipeline injection
"""
import warnings; warnings.filterwarnings('ignore')
import pandas as pd
import numpy as np
import io
import streamlit as st
from datetime import datetime
from feature_engineering import engineer, LEAKAGE_SQ, LEAKAGE_PP, TARGETS
from date_utils import parse_dates_robust
from column_mapper import map_columns

# ── Column definitions ─────────────────────────────────────────────────────

# Sinter Quality sheet expected columns
SQ_DATE_COL   = 'DATE'
SQ_REQUIRED   = ['DATE']
SQ_NUMERIC    = ['%T(Fe)','%FeO','%CaO','%MgO','%SiO2','%Al2O3',
                 '%K2O','% P','Avl. lime','Basicity  (B2)','MgO/Al2O3',
                 'TI','RDI','RI']
SQ_TARGETS    = ['TI','RDI','RI']

# Process Parameters sheet expected columns
PP_DATE_COL   = 'Date'
PP_REQUIRED   = ['Date']
PP_NUMERIC    = ['Feed','Moisture','Furnace Temp.','BF gas flow','WGF Speed',
                 'Machine speed','Cooler Speed','BTP Temperature','WB 1 Pressure',
                 'ESP Inlet Temp.','ESP Inlet Press.','Cooler Underneath Pressure',
                 'Product Temp','Old ESP Opacity','%T(Fe)','%FeO','%CaO','%MgO',
                 '%SiO2','%Al2O3','%K2O','% P','Avl. lime','Basicity(B2)',
                 'Carbon Rate','Production',
                 'WF#11','WF#12','WF#13','WF#14','Inhouse','OMC+Tensa','NMDC',
                 'Gadchiroli','Nmdc bf','Concentrate','Kasia from BF','Kasia hg',
                 'ROIDA','vedanta','PTA Fines','MGM FINES']

# Leakage columns to always drop
LEAKAGE_DROP  = ['%+40MM','%-40+25MM','%-25+10MM','%-10+5MM','%-5MM','%MPS',
                 '%6.3MM(TI)','%-0.5MM','G.I.','B.I.','Fines Generation',
                 '%40+25MM','%25+10MM','%10+5MM','%5 MM']

BF_TARGETS_VAL = {'TI':78.0,'RDI':25.0,'RI':68.0}

# ── Validation helpers ─────────────────────────────────────────────────────

def _check_result(ok, msg, detail=''):
    return {'ok': ok, 'msg': msg, 'detail': detail}


def validate_dataframe(df, date_col, required_numerics, dataset_label):
    """
    Run all validation checks on a freshly parsed DataFrame.
    Returns list of result dicts.
    """
    results = []

    # 1. Empty check
    if df is None or len(df) == 0:
        results.append(_check_result(False, f'{dataset_label}: File is empty or could not be parsed.'))
        return results
    results.append(_check_result(True, f'{dataset_label}: File parsed — {len(df):,} rows, {df.shape[1]} columns.'))

    # 2. Date column present
    if date_col not in df.columns:
        similar = [c for c in df.columns if 'date' in c.lower() or 'time' in c.lower()]
        detail  = f'Did you mean: {similar}?' if similar else f'Found columns: {df.columns.tolist()[:8]}'
        results.append(_check_result(False, f'{dataset_label}: Date column "{date_col}" not found.', detail))
        return results
    results.append(_check_result(True, f'{dataset_label}: Date column "{date_col}" found.'))

    # 3. Date parseable
    parsed_dates = parse_dates_robust(df[date_col])
    n_bad_dates  = parsed_dates.isna().sum()
    if n_bad_dates == len(df):
        results.append(_check_result(False, f'{dataset_label}: Cannot parse any dates in "{date_col}".',
                                     f'Sample values: {df[date_col].head(3).tolist()}'))
        return results
    elif n_bad_dates > 0:
        results.append(_check_result(True,
            f'{dataset_label}: {n_bad_dates} unparseable date rows will be dropped.',
            f'Valid dates: {len(df)-n_bad_dates}'))
    else:
        results.append(_check_result(True,
            f'{dataset_label}: All {len(df)} dates parsed successfully.',
            f'Range: {parsed_dates.min().strftime("%d %b %Y")} → {parsed_dates.max().strftime("%d %b %Y")}'))

    # 4. Numeric columns
    found_num   = [c for c in required_numerics if c in df.columns]
    missing_num = [c for c in required_numerics if c not in df.columns]
    if len(found_num) == 0:
        results.append(_check_result(False,
            f'{dataset_label}: None of the expected numeric columns found.',
            f'Expected sample: {required_numerics[:5]}'))
    elif missing_num:
        results.append(_check_result(True,
            f'{dataset_label}: {len(found_num)} of {len(required_numerics)} numeric columns found.',
            f'Missing (will be skipped): {missing_num[:6]}'))
    else:
        results.append(_check_result(True,
            f'{dataset_label}: All {len(found_num)} numeric columns found.'))

    # 5. Missing value report
    if found_num:
        null_pct = df[found_num].isna().mean() * 100
        high_null = null_pct[null_pct > 50].index.tolist()
        if high_null:
            results.append(_check_result(True,
                f'{dataset_label}: {len(high_null)} columns have >50% missing values — will still aggregate.',
                f'Columns: {high_null[:4]}'))
        else:
            avg_null = null_pct.mean()
            results.append(_check_result(True,
                f'{dataset_label}: Average missing data {avg_null:.1f}% — acceptable.'))

    # 6. Duplicate dates warning
    dup_dates = parsed_dates.dropna().duplicated().sum()
    if dup_dates > 0:
        results.append(_check_result(True,
            f'{dataset_label}: {dup_dates} duplicate date entries → will be averaged together.',
            'This is normal for sub-daily data.'))
    else:
        results.append(_check_result(True,
            f'{dataset_label}: No duplicate dates — one row per day.'))

    return results


# ── Core aggregation ────────────────────────────────────────────────────────


def _detect_shutdown_dates(df, date_col):
    """
    Detect shutdown days — any row where any column contains
    'SD', 'SD/', 'Shutdown', 'SHUTDOWN', 'shutdown' as text value.
    Returns set of datetime.date objects.
    """
    shutdown_keywords = ['sd', 'sd/', 'shutdown', 'shut down', 'shut-down', 'plant off', 'no production']
    shutdown_mask = pd.Series(False, index=df.index)
    for col in df.columns:
        if col == date_col:
            continue
        try:
            str_col = df[col].astype(str).str.strip().str.lower()
            col_mask = str_col.isin(shutdown_keywords)
            # Also catch partial matches like "SD/Shutdown", "SD - Furnace Off"
            col_mask |= str_col.str.startswith('sd') & ~str_col.str.match(r'^\d')
            shutdown_mask |= col_mask
        except Exception:
            pass
    if shutdown_mask.any():
        return set(df.loc[shutdown_mask, date_col].dt.normalize().dropna().unique())
    return set()

def aggregate_to_daily(df, date_col, numeric_cols, dataset_label):
    """
    Parse dates, coerce numerics, group by date, compute daily mean.
    Returns (aggregated_df, stats_dict)
    """
    df = df.copy()

    # Parse dates
    df[date_col] = parse_dates_robust(df[date_col])
    df = df.dropna(subset=[date_col])
    df[date_col] = df[date_col].dt.normalize()   # strip time component → date only

    # Drop leakage columns
    df.drop(columns=[c for c in LEAKAGE_DROP if c in df.columns], inplace=True, errors='ignore')

    # Coerce all numeric cols
    avail_num = [c for c in numeric_cols if c in df.columns]
    for c in avail_num:
        df[c] = pd.to_numeric(df[c], errors='coerce')

    # Remove extreme RDI outliers if present (same as training pipeline)
    if 'RDI' in df.columns:  # RDI < 35 — remove extreme outliers same as training
        df = df[df['RDI'].isna() | (df['RDI'] < 35)]

    # Detect Shutdown dates — any row with 'SD', 'SD/', 'Shutdown' in ANY column
    shutdown_dates = _detect_shutdown_dates(df, date_col)

    # Group by date → daily mean
    agg = df.groupby(date_col)[avail_num].mean().reset_index()
    agg = agg.sort_values(date_col).reset_index(drop=True)

    # Tag shutdown dates
    agg['Shutdown'] = agg[date_col].isin(shutdown_dates).map({True:'🔴 Shutdown', False:''})

    stats = {
        'input_rows':    len(df),
        'output_days':   len(agg),
        'date_range':    (agg[date_col].min(), agg[date_col].max()),
        'cols_agg':      avail_num,
        'nulls_after':   agg[avail_num].isna().sum().to_dict(),
        'shutdown_dates': sorted([d.strftime('%d %b %Y') for d in shutdown_dates]),
        'n_shutdowns':   len(shutdown_dates),
    }
    return agg, stats


def _auto_map(raw_df, date_col, numeric_cols):
    """Fuzzy-map slightly different column headers onto the pipeline's
    expected schema. Returns (mapped_df, {actual: expected} for anything
    that actually changed — empty dict if every header already matched).
    """
    expected = [date_col] + numeric_cols
    full_map = map_columns(raw_df.columns.tolist(), expected)
    applied  = {a: e for a, e in full_map.items() if a != e}
    return (raw_df.rename(columns=applied), applied) if applied else (raw_df, {})


def process_sq_upload(raw_df):
    """Full SQ pipeline: auto-map columns → validate → aggregate → engineer features."""
    mapped, mapping = _auto_map(raw_df, SQ_DATE_COL, SQ_NUMERIC)
    agg, stats = aggregate_to_daily(mapped, SQ_DATE_COL, SQ_NUMERIC, 'Sinter Quality')
    sq_fe = engineer(agg, is_pp=False)
    stats['column_mapping'] = mapping
    return agg, sq_fe, stats


def process_pp_upload(raw_df):
    """Full PP pipeline: auto-map columns → validate → aggregate → engineer features."""
    mapped, mapping = _auto_map(raw_df, PP_DATE_COL, PP_NUMERIC)
    agg, stats = aggregate_to_daily(mapped, PP_DATE_COL, PP_NUMERIC, 'Process Parameters')
    pp_fe = engineer(agg, is_pp=True)
    stats['column_mapping'] = mapping
    return agg, pp_fe, stats


# ── Preview helpers ─────────────────────────────────────────────────────────

def quality_summary_html(agg_df):
    """Generate a quick quality summary card for SQ data."""
    rows = []
    for t in SQ_TARGETS:
        if t not in agg_df.columns:
            continue
        s   = agg_df[t].dropna()
        tgt = BF_TARGETS_VAL[t]
        direction = '>=' if t != 'RDI' else '<='
        in_pct = ((s >= tgt).mean() if direction == '>=' else (s <= tgt).mean()) * 100
        ok  = in_pct >= 80
        col = '#22c55e' if ok else '#ef4444'
        rows.append(f"""
        <div style="flex:1;background:#060d1a;border:1px solid #1e3a5f;
                    border-top:3px solid {col};border-radius:8px;padding:12px;text-align:center">
          <div style="font-size:0.68rem;color:#64748b;margin-bottom:4px">{t}</div>
          <div style="font-size:1.3rem;font-weight:700;color:{col};
                      font-family:'JetBrains Mono',monospace">{s.mean():.3f}</div>
          <div style="font-size:0.65rem;color:#64748b;margin-top:4px">
            σ={s.std():.3f} &nbsp;|&nbsp; In-target: {in_pct:.0f}%
          </div>
        </div>""")
    return f'<div style="display:flex;gap:10px;margin:10px 0">{"".join(rows)}</div>'


# ── Main render ─────────────────────────────────────────────────────────────

def render(existing_sq, existing_pp, sq_fe_existing, art):
    """
    Main render for the Data Ingestion page.
    existing_sq / existing_pp = the currently loaded DataFrames from artifacts.
    Returns (updated_sq, updated_pp, updated_sq_fe) or originals if no upload.
    """

    # ── CSS additions ──────────────────────────────────────────────────────
    st.markdown("""
    <style>
    .upload-zone {
        background: linear-gradient(135deg,#0a1628,#0d1f3c);
        border: 2px dashed #1e4080;
        border-radius:12px; padding:24px 20px;
        text-align:center; transition:all 0.2s;
    }
    .upload-zone:hover { border-color:#2563eb; }
    .check-ok   { color:#22c55e; font-size:0.78rem; margin:3px 0; }
    .check-warn { color:#f59e0b; font-size:0.78rem; margin:3px 0; }
    .check-fail { color:#ef4444; font-size:0.78rem; margin:3px 0; }
    .stat-pill  {
        display:inline-block; background:#0a1628;
        border:1px solid #1e3a5f; border-radius:20px;
        padding:3px 12px; font-size:0.7rem; color:#94b4d4; margin:3px;
    }
    .section-badge {
        background:#0d1f3c; border-left:3px solid #2563eb;
        padding:8px 14px; font-size:0.78rem; font-weight:700;
        color:#7eb3f7; letter-spacing:0.5px;
        border-radius:0 6px 6px 0; margin:14px 0 10px 0;
    }
    .merge-badge {
        background:#052e16; border:1px solid #22c55e;
        border-radius:8px; padding:12px 16px;
        font-size:0.8rem; color:#86efac; margin:10px 0;
    }
    </style>""", unsafe_allow_html=True)

    # ── Page header ────────────────────────────────────────────────────────
    st.markdown("""
    <div style="background:linear-gradient(135deg,#0a1628,#0f2347);
                border:1px solid #1e3a5f;border-radius:12px;padding:20px 24px;margin-bottom:18px">
      <div style="font-size:1.1rem;font-weight:700;color:#e2e8f0;margin-bottom:4px">
        📂 Automated Excel Data Aggregation & Processing
      </div>
      <div style="font-size:0.78rem;color:#64748b;line-height:1.6">
        Upload raw Excel files from your sinter plant.
        The system automatically detects columns, aggregates to daily averages,
        applies the feature engineering pipeline, and merges with existing data
        to update the full dashboard in real time.
      </div>
    </div>""", unsafe_allow_html=True)

    # ── How it works ───────────────────────────────────────────────────────
    with st.expander('ℹ️ How This Works — Click to Expand', expanded=False):
        st.markdown("""
**Step 1 — Upload** your raw Excel file (.xlsx or .xls). It can be sub-daily (hourly/shift-wise) or already daily.

**Step 2 — Auto Detection** scans for the date column (`DATE` for Sinter Quality, `Date` for Process Parameters), validates all expected columns, and reports any missing or problematic columns.

**Step 3 — Daily Aggregation** groups all rows by date and computes the **mean** of every numeric column. Sub-daily readings (hourly, shift-wise) are automatically collapsed into one row per day.

**Step 4 — Pipeline Processing** applies the same feature engineering used during model training (Basicity ratios, MgO/Al₂O₃, rolling averages, lagged features, interaction terms).

**Step 5 — Merge & Update** combines the uploaded data with existing data and updates all dashboard pages (trends, SHAP, predictions, BF suitability) with the new data.

| Dataset | Date Column | Key Columns |
|---|---|---|
| Sinter Quality | `DATE` | TI, RDI, RI, %FeO, %CaO, %MgO, %SiO₂, %Al₂O₃, Basicity |
| Process Parameters | `Date` | Machine speed, BTP Temp, WGF Speed, Feed, Moisture, ESP Temp/Press |
        """)

    # ── Upload tabs ────────────────────────────────────────────────────────
    tab_sq, tab_pp, tab_both = st.tabs([
        '🧪 Sinter Quality Data',
        '⚙️ Process Parameters Data',
        '📦 Upload Both Together',
    ])

    sq_result = pp_result = None  # will hold (agg_df, fe_df, stats) tuples

    # ══════════════════════════════════════════════════════════════════
    # TAB 1 — SINTER QUALITY
    # ══════════════════════════════════════════════════════════════════
    with tab_sq:
        st.markdown('<div class="section-badge">📤 Upload Sinter Quality Excel File</div>',
                    unsafe_allow_html=True)
        st.markdown(f"""
        <div style="font-size:0.74rem;color:#64748b;margin-bottom:12px">
          Expected date column: <code style="color:#7eb3f7">DATE</code> &nbsp;·&nbsp;
          Key columns: <code style="color:#7eb3f7">TI, RDI, RI, %FeO, %CaO, %MgO, %SiO₂, %Al₂O₃, Basicity&nbsp;(B2)</code>
        </div>""", unsafe_allow_html=True)

        sq_file = st.file_uploader(
            'Drop Sinter Quality Excel here',
            type=['xlsx','xls'],
            key='sq_uploader',
            label_visibility='collapsed',
        )

        if sq_file:
            with st.spinner('Parsing and validating Sinter Quality data...'):
                try:
                    raw_sq = _parse_excel(sq_file)
                except Exception as e:
                    st.error(f'❌ Cannot read file: {e}')
                    raw_sq = None

            if raw_sq is not None:
                _render_validation(raw_sq, SQ_DATE_COL, SQ_NUMERIC, 'Sinter Quality')

                all_ok = SQ_DATE_COL in raw_sq.columns
                if all_ok:
                    if st.button('🚀 Process & Aggregate Sinter Quality Data',
                                 key='btn_sq', use_container_width=True):
                        with st.spinner('Aggregating to daily averages and engineering features...'):
                            try:
                                agg, fe, stats = process_sq_upload(raw_sq)
                                st.session_state['sq_upload_result'] = (agg, fe, stats)
                                st.success(f'✅ Processing complete — {stats["output_days"]} daily records created.')
                            except Exception as e:
                                st.error(f'❌ Processing failed: {e}')

            if 'sq_upload_result' in st.session_state:
                agg, fe, stats = st.session_state['sq_upload_result']
                _render_sq_results(agg, fe, stats, existing_sq)

    # ══════════════════════════════════════════════════════════════════
    # TAB 2 — PROCESS PARAMETERS
    # ══════════════════════════════════════════════════════════════════
    with tab_pp:
        st.markdown('<div class="section-badge">📤 Upload Process Parameters Excel File</div>',
                    unsafe_allow_html=True)
        st.markdown(f"""
        <div style="font-size:0.74rem;color:#64748b;margin-bottom:12px">
          Expected date column: <code style="color:#7eb3f7">Date</code> &nbsp;·&nbsp;
          Key columns: <code style="color:#7eb3f7">Machine speed, BTP Temperature, WGF Speed, Feed, Moisture, ESP Inlet Temp., %FeO, Basicity(B2)</code>
        </div>""", unsafe_allow_html=True)

        pp_file = st.file_uploader(
            'Drop Process Parameters Excel here',
            type=['xlsx','xls'],
            key='pp_uploader',
            label_visibility='collapsed',
        )

        if pp_file:
            with st.spinner('Parsing and validating Process Parameters data...'):
                try:
                    raw_pp = _parse_excel(pp_file)
                except Exception as e:
                    st.error(f'❌ Cannot read file: {e}')
                    raw_pp = None

            if raw_pp is not None:
                _render_validation(raw_pp, PP_DATE_COL, PP_NUMERIC, 'Process Parameters')

                all_ok = PP_DATE_COL in raw_pp.columns
                if all_ok:
                    if st.button('🚀 Process & Aggregate Process Parameter Data',
                                 key='btn_pp', use_container_width=True):
                        with st.spinner('Aggregating to daily averages and engineering features...'):
                            try:
                                agg, fe, stats = process_pp_upload(raw_pp)
                                st.session_state['pp_upload_result'] = (agg, fe, stats)
                                st.success(f'✅ Processing complete — {stats["output_days"]} daily records created.')
                            except Exception as e:
                                st.error(f'❌ Processing failed: {e}')

            if 'pp_upload_result' in st.session_state:
                agg, fe, stats = st.session_state['pp_upload_result']
                _render_pp_results(agg, fe, stats, existing_pp)

    # ══════════════════════════════════════════════════════════════════
    # TAB 3 — BOTH FILES (multi-sheet or separate uploads)
    # ══════════════════════════════════════════════════════════════════
    with tab_both:
        st.markdown('<div class="section-badge">📤 Upload Combined Excel (Multi-Sheet)</div>',
                    unsafe_allow_html=True)
        st.markdown("""
        <div style="font-size:0.74rem;color:#64748b;margin-bottom:12px">
          Upload a single Excel file with multiple sheets — the system auto-detects
          which sheet is Sinter Quality and which is Process Parameters.
        </div>""", unsafe_allow_html=True)

        combined_file = st.file_uploader(
            'Drop Combined Excel here',
            type=['xlsx','xls'],
            key='combined_uploader',
            label_visibility='collapsed',
        )

        if combined_file:
            with st.spinner('Reading sheets...'):
                try:
                    sheets = _detect_sheets(combined_file)
                except Exception as e:
                    st.error(f'❌ Cannot read file: {e}')
                    sheets = {}

            if sheets:
                st.markdown(f'<div style="font-size:0.74rem;color:#94b4d4;margin-bottom:8px">'
                            f'📋 Detected {len(sheets)} sheet(s): '
                            f'{", ".join([f"<b>{k}</b> ({v} rows)" for k,v in [(s,len(d)) for s,d in sheets.items()]])}'
                            f'</div>', unsafe_allow_html=True)

                sq_sheet = _auto_detect_sq_sheet(sheets)
                pp_sheet = _auto_detect_pp_sheet(sheets)

                c1, c2 = st.columns(2)
                with c1:
                    sel_sq = st.selectbox('Sheet for Sinter Quality',
                                          ['(none)'] + list(sheets.keys()),
                                          index=list(sheets.keys()).index(sq_sheet)+1 if sq_sheet else 0,
                                          key='sel_sq_sheet')
                with c2:
                    sel_pp = st.selectbox('Sheet for Process Parameters',
                                          ['(none)'] + list(sheets.keys()),
                                          index=list(sheets.keys()).index(pp_sheet)+1 if pp_sheet else 0,
                                          key='sel_pp_sheet')

                if st.button('🚀 Process Both Sheets', key='btn_both', use_container_width=True):
                    with st.spinner('Processing both datasets...'):
                        success_sq = success_pp = False
                        if sel_sq != '(none)':
                            try:
                                agg, fe, stats = process_sq_upload(sheets[sel_sq])
                                st.session_state['sq_upload_result'] = (agg, fe, stats)
                                success_sq = True
                                st.success(f'✅ Sinter Quality: {stats["output_days"]} daily records.')
                            except Exception as e:
                                st.error(f'❌ Sinter Quality failed: {e}')
                        if sel_pp != '(none)':
                            try:
                                agg, fe, stats = process_pp_upload(sheets[sel_pp])
                                st.session_state['pp_upload_result'] = (agg, fe, stats)
                                success_pp = True
                                st.success(f'✅ Process Parameters: {stats["output_days"]} daily records.')
                            except Exception as e:
                                st.error(f'❌ Process Parameters failed: {e}')

    # ── Merge control ──────────────────────────────────────────────────────
    has_sq = 'sq_upload_result' in st.session_state
    has_pp = 'pp_upload_result' in st.session_state

    if has_sq or has_pp:
        st.markdown('---')
        st.markdown('<div class="section-badge">🔄 MERGE UPLOADED DATA INTO DASHBOARD</div>',
                    unsafe_allow_html=True)

        merge_mode = st.radio(
            'Merge strategy',
            ['Append (add new dates to existing data)',
             'Replace (use uploaded data only)',
             'Update (overwrite overlapping dates, keep rest)'],
            horizontal=True,
            label_visibility='collapsed',
        )

        col_m1, col_m2 = st.columns(2)
        if has_sq:
            agg_sq, fe_sq, stats_sq = st.session_state['sq_upload_result']
            col_m1.markdown(
                f'<div class="merge-badge">🧪 <b>Sinter Quality ready</b><br>'
                f'{stats_sq["output_days"]} daily records &nbsp;·&nbsp; '
                f'{stats_sq["date_range"][0].strftime("%d %b %Y")} → '
                f'{stats_sq["date_range"][1].strftime("%d %b %Y")}</div>',
                unsafe_allow_html=True)
        if has_pp:
            agg_pp, fe_pp, stats_pp = st.session_state['pp_upload_result']
            col_m2.markdown(
                f'<div class="merge-badge">⚙️ <b>Process Parameters ready</b><br>'
                f'{stats_pp["output_days"]} daily records &nbsp;·&nbsp; '
                f'{stats_pp["date_range"][0].strftime("%d %b %Y")} → '
                f'{stats_pp["date_range"][1].strftime("%d %b %Y")}</div>',
                unsafe_allow_html=True)

        if st.button('✅ Apply & Update Dashboard', key='btn_merge',
                     use_container_width=True, type='primary'):
            mode_key = ('append'  if 'Append'  in merge_mode else
                        'replace' if 'Replace' in merge_mode else 'update')
            updated_sq = existing_sq.copy()
            updated_pp = existing_pp.copy()

            if has_sq:
                agg_sq, fe_sq, _ = st.session_state['sq_upload_result']
                updated_sq = _merge(existing_sq, agg_sq, SQ_DATE_COL, mode_key)

            if has_pp:
                agg_pp, fe_pp, _ = st.session_state['pp_upload_result']
                updated_pp = _merge(existing_pp, agg_pp, PP_DATE_COL, mode_key)

            # Re-engineer features on merged SQ
            updated_sq_fe = engineer(updated_sq, is_pp=False)

            # Persist into session state so all other pages pick it up
            st.session_state['live_sq']    = updated_sq
            st.session_state['live_pp']    = updated_pp
            st.session_state['live_sq_fe'] = updated_sq_fe
            st.session_state['data_updated'] = True

            st.success(
                f'✅ Dashboard updated! '
                f'Sinter Quality: {len(updated_sq):,} records | '
                f'Process Params: {len(updated_pp):,} records. '
                f'Navigate to any page to see updated data.')

        if st.session_state.get('data_updated'):
            st.markdown("""
            <div style="background:#0f2347;border:1px solid #2563eb;border-radius:8px;
                        padding:12px 16px;font-size:0.78rem;color:#7eb3f7">
              🟢 <b>Dashboard is using uploaded data.</b>
              All pages (Quality Monitoring, Process Monitoring, Predictions, SHAP, BF Suitability)
              are now reflecting the merged dataset.
              &nbsp;<a href="#" style="color:#f97316">Click Reset below to restore original data.</a>
            </div>""", unsafe_allow_html=True)

            if st.button('🔄 Reset to Original Data', key='btn_reset'):
                for key in ['sq_upload_result','pp_upload_result',
                            'live_sq','live_pp','live_sq_fe','data_updated']:
                    st.session_state.pop(key, None)
                st.success('Restored original data. Reloading...')
                st.rerun()

    # ── Return live data to app.py ─────────────────────────────────────────
    if st.session_state.get('data_updated'):
        return (st.session_state['live_sq'],
                st.session_state['live_pp'],
                st.session_state['live_sq_fe'])
    return existing_sq, existing_pp, sq_fe_existing


# ── Internal helpers ────────────────────────────────────────────────────────

def _parse_excel(file_obj):
    """Read uploaded file object into DataFrame, try multiple header rows."""
    raw = file_obj.read()
    buf = io.BytesIO(raw)
    # Try standard read first
    try:
        df = pd.read_excel(buf, header=0)
        if df.shape[1] > 1:
            return df
    except Exception:
        pass
    # Try with header=1 (some files have merged title row)
    buf.seek(0)
    try:
        df = pd.read_excel(buf, header=1)
        return df
    except Exception as e:
        raise ValueError(f'Could not parse Excel: {e}')


def _detect_sheets(file_obj):
    """Return dict of {sheet_name: DataFrame} for all sheets."""
    raw  = file_obj.read()
    buf  = io.BytesIO(raw)
    xl   = pd.ExcelFile(buf)
    out  = {}
    for name in xl.sheet_names:
        try:
            df = pd.read_excel(buf, sheet_name=name, header=0)
            if len(df) > 0:
                out[name] = df
        except Exception:
            pass
    return out


def _auto_detect_sq_sheet(sheets):
    """Heuristic: SQ sheet has TI, RDI, RI columns."""
    for name, df in sheets.items():
        cols_lower = [c.lower() for c in df.columns]
        if any('ti' == c for c in cols_lower) and 'rdi' in cols_lower:
            return name
        if 'date' in cols_lower and any('%feo' in c for c in [c.lower() for c in df.columns]):
            return name
    return list(sheets.keys())[0] if sheets else None


def _auto_detect_pp_sheet(sheets):
    """Heuristic: PP sheet has Machine speed / BTP Temperature."""
    for name, df in sheets.items():
        cols_lower = [c.lower() for c in df.columns]
        if 'machine speed' in ' '.join(cols_lower) or 'btp' in ' '.join(cols_lower):
            return name
    keys = list(sheets.keys())
    return keys[1] if len(keys) > 1 else keys[0] if keys else None


def _render_validation(df, date_col, num_cols, label):
    """Show expandable validation results with colour-coded checks."""
    results = validate_dataframe(df, date_col, num_cols, label)
    all_ok  = all(r['ok'] for r in results)
    icon    = '✅' if all_ok else '⚠️'

    with st.expander(f'{icon} Validation Report — {label}', expanded=not all_ok):
        for r in results:
            css = 'check-ok' if r['ok'] else 'check-fail'
            icon2 = '✅' if r['ok'] else '❌'
            detail_html = (f'<span style="color:#475569;font-size:0.7rem"> — {r["detail"]}</span>'
                           if r['detail'] else '')
            st.markdown(f'<div class="{css}">{icon2} {r["msg"]}{detail_html}</div>',
                        unsafe_allow_html=True)

        # Column mapping table
        st.markdown('<div style="font-size:0.72rem;color:#64748b;margin-top:10px;margin-bottom:4px">'
                    'Column Detection:</div>', unsafe_allow_html=True)
        found   = [c for c in num_cols if c in df.columns]
        missing = [c for c in num_cols if c not in df.columns]
        col_df  = pd.DataFrame({
            'Column': found + missing,
            'Status': ['✅ Found'] * len(found) + ['❌ Missing'] * len(missing),
            'Non-null %': [
                f'{df[c].notna().mean()*100:.0f}%' for c in found
            ] + ['—'] * len(missing),
        })
        st.dataframe(col_df, use_container_width=True, height=200, hide_index=True)



def _to_excel_bytes(df):
    """Convert DataFrame to Excel bytes for st.download_button."""
    import io as _io
    from openpyxl.styles import Font, PatternFill, Alignment
    buf = _io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Daily Processed')
        ws = writer.sheets['Daily Processed']
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF', name='Calibri')
            cell.fill = PatternFill('solid', start_color='1A2744')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Highlight Shutdown column red
        if 'Shutdown' in df.columns:
            sd_col_idx = df.columns.get_loc('Shutdown') + 1
            for row in ws.iter_rows(min_row=2, min_col=sd_col_idx, max_col=sd_col_idx):
                for cell in row:
                    if cell.value and 'Shutdown' in str(cell.value):
                        cell.fill = PatternFill('solid', start_color='7F1D1D')
                        cell.font = Font(bold=True, color='FCA5A5', name='Calibri')
        # Auto-width
        for col in ws.columns:
            from openpyxl.utils import get_column_letter
            ml = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ml+2,8),25)
    buf.seek(0)
    return buf.getvalue()

def _render_sq_results(agg, fe, stats, existing_sq):
    """Show SQ aggregation results, preview, quality summary."""
    st.markdown('<div class="section-badge">📊 Aggregation Results</div>', unsafe_allow_html=True)

    mapping = stats.get('column_mapping') or {}
    if mapping:
        pairs = ' &nbsp;·&nbsp; '.join(f'<code>{a}</code> → <code>{e}</code>' for a, e in mapping.items())
        st.markdown(
            f'<div style="font-size:0.72rem;color:#94b4d4;background:#0a1628;'
            f'border:1px solid #1e3a5f;border-radius:6px;padding:8px 12px;margin-bottom:8px">'
            f'🔧 <b>Auto-mapped columns:</b> {pairs}</div>', unsafe_allow_html=True)

    p1,p2,p3,p4 = st.columns(4)
    p1.metric('Input Rows',    f'{stats["input_rows"]:,}')
    p2.metric('Daily Records', f'{stats["output_days"]:,}')
    p3.metric('Date From',     stats['date_range'][0].strftime('%d %b %Y'))
    p4.metric('Date To',       stats['date_range'][1].strftime('%d %b %Y'))

    # Quality summary cards
    if any(t in agg.columns for t in SQ_TARGETS):
        st.markdown(quality_summary_html(agg), unsafe_allow_html=True)

    # Daily preview table
    st.markdown('<div style="font-size:0.74rem;color:#64748b;margin:10px 0 4px">Preview — Daily Aggregated Data (first 10 rows)</div>',
                unsafe_allow_html=True)
    preview_cols = [SQ_DATE_COL] + [c for c in SQ_NUMERIC if c in agg.columns][:8]
    st.dataframe(agg[preview_cols].head(10).round(3),
                 use_container_width=True, hide_index=True)

    # Overlap with existing data
    existing_dates = set(parse_dates_robust(existing_sq[SQ_DATE_COL]).dropna())
    new_dates      = set(agg[SQ_DATE_COL])
    overlap        = len(existing_dates & new_dates)
    new_only       = len(new_dates - existing_dates)
    st.markdown(
        f'<div style="font-size:0.74rem;color:#94b4d4;background:#0a1628;'
        f'border:1px solid #1e3a5f;border-radius:6px;padding:10px 14px;margin-top:8px">'
        f'📅 <b>Overlap with existing data:</b> {overlap} common dates &nbsp;·&nbsp; '
        f'<span style="color:#22c55e">{new_only} new dates</span> to be added</div>',
        unsafe_allow_html=True)

    # Download aggregated
    # Show shutdown summary
    n_sd = stats.get('n_shutdowns', 0)
    if n_sd > 0:
        sd_dates = ', '.join(stats.get('shutdown_dates', [])[:5])
        st.markdown(
            f'<div style="background:#1c0202;border:1px solid #ef4444;border-radius:6px;'
            f'padding:10px 14px;margin:8px 0;font-size:0.75rem;">'
            f'🔴 <b style="color:#ef4444">{n_sd} Shutdown Day(s) Detected</b>'
            f'<span style="color:#94b4d4"> — tagged in output file: {sd_dates}'
            f'{"..." if n_sd > 5 else ""}</span></div>',
            unsafe_allow_html=True)

    # Excel download (includes Shutdown column)
    excel_buf = _to_excel_bytes(agg)
    st.download_button('⬇️ Download Processed Daily Average (.xlsx)',
                       data=excel_buf,
                       file_name='sq_daily_processed.xlsx',
                       mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                       key='dl_sq')


def _render_pp_results(agg, fe, stats, existing_pp):
    """Show PP aggregation results and preview."""
    st.markdown('<div class="section-badge">📊 Aggregation Results</div>', unsafe_allow_html=True)

    mapping = stats.get('column_mapping') or {}
    if mapping:
        pairs = ' &nbsp;·&nbsp; '.join(f'<code>{a}</code> → <code>{e}</code>' for a, e in mapping.items())
        st.markdown(
            f'<div style="font-size:0.72rem;color:#94b4d4;background:#0a1628;'
            f'border:1px solid #1e3a5f;border-radius:6px;padding:8px 12px;margin-bottom:8px">'
            f'🔧 <b>Auto-mapped columns:</b> {pairs}</div>', unsafe_allow_html=True)

    p1,p2,p3,p4 = st.columns(4)
    p1.metric('Input Rows',    f'{stats["input_rows"]:,}')
    p2.metric('Daily Records', f'{stats["output_days"]:,}')
    p3.metric('Date From',     stats['date_range'][0].strftime('%d %b %Y'))
    p4.metric('Date To',       stats['date_range'][1].strftime('%d %b %Y'))

    preview_cols = [PP_DATE_COL] + [c for c in PP_NUMERIC if c in agg.columns][:8]
    st.markdown('<div style="font-size:0.74rem;color:#64748b;margin:10px 0 4px">Preview — Daily Aggregated Data (first 10 rows)</div>',
                unsafe_allow_html=True)
    st.dataframe(agg[preview_cols].head(10).round(3),
                 use_container_width=True, hide_index=True)

    existing_dates = set(parse_dates_robust(existing_pp[PP_DATE_COL]).dropna())
    new_dates      = set(agg[PP_DATE_COL])
    overlap        = len(existing_dates & new_dates)
    new_only       = len(new_dates - existing_dates)
    st.markdown(
        f'<div style="font-size:0.74rem;color:#94b4d4;background:#0a1628;'
        f'border:1px solid #1e3a5f;border-radius:6px;padding:10px 14px;margin-top:8px">'
        f'📅 <b>Overlap with existing data:</b> {overlap} common dates &nbsp;·&nbsp; '
        f'<span style="color:#22c55e">{new_only} new dates</span> to be added</div>',
        unsafe_allow_html=True)

    n_sd = stats.get('n_shutdowns', 0)
    if n_sd > 0:
        sd_dates = ', '.join(stats.get('shutdown_dates', [])[:5])
        st.markdown(
            f'<div style="background:#1c0202;border:1px solid #ef4444;border-radius:6px;'
            f'padding:10px 14px;margin:8px 0;font-size:0.75rem;">'
            f'🔴 <b style="color:#ef4444">{n_sd} Shutdown Day(s) Detected</b>'
            f'<span style="color:#94b4d4"> — tagged in output: {sd_dates}</span></div>',
            unsafe_allow_html=True)

    excel_buf = _to_excel_bytes(agg)
    st.download_button('⬇️ Download Processed Daily Average (.xlsx)',
                       data=excel_buf,
                       file_name='pp_daily_processed.xlsx',
                       mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                       key='dl_pp')


def _merge(existing_df, new_df, date_col, mode):
    """Merge new data into existing according to selected strategy."""
    existing = existing_df.copy()
    new      = new_df.copy()
    existing[date_col] = parse_dates_robust(existing[date_col])
    new[date_col]      = parse_dates_robust(new[date_col])

    if mode == 'replace':
        return new.sort_values(date_col).reset_index(drop=True)

    elif mode == 'append':
        existing_dates = set(existing[date_col].dropna())
        new_only = new[~new[date_col].isin(existing_dates)]
        merged   = pd.concat([existing, new_only], ignore_index=True)
        return merged.sort_values(date_col).reset_index(drop=True)

    elif mode == 'update':
        # Remove existing rows whose date appears in new data, then add all new rows
        new_dates = set(new[date_col].dropna())
        existing_kept = existing[~existing[date_col].isin(new_dates)]
        merged = pd.concat([existing_kept, new], ignore_index=True)
        return merged.sort_values(date_col).reset_index(drop=True)

    return existing
